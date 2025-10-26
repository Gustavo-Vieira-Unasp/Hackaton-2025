from openpyxl import load_workbook, Workbook
import json

# Função: Excel para JSON
def excel_para_json_openpyxl(caminho_excel: str) -> str:
    wb = load_workbook(caminho_excel)
    ws = wb.active

    colunas = [cell.value for cell in ws[1]]  # primeira linha como cabeçalho
    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        registro = {colunas[i]: row[i] for i in range(len(colunas))}
        dados.append(registro)

    return json.dumps(dados, ensure_ascii=False, indent=4)

# Função: JSON para Excel
def json_para_excel_openpyxl(caminho_json: str, caminho_excel: str):
    with open(caminho_json, "r", encoding="utf-8") as f:
        dados = json.load(f)

    wb = Workbook()
    ws = wb.active

    if len(dados) == 0:
        wb.save(caminho_excel)
        return

    # Escrevendo cabeçalho
    ws.append(list(dados[0].keys()))

    # Escrevendo dados
    for registro in dados:
        ws.append(list(registro.values()))

    wb.save(caminho_excel)

# ===== Exemplo de uso =====
if __name__ == "__main__":
    # Excel para JSON
    json_str = excel_para_json_openpyxl("registro.xlsx")
    with open("registro_openpyxl.json", "w", encoding="utf-8") as f:
        f.write(json_str)
    print("Excel convertido para JSON com openpyxl: registro_openpyxl.json")

    # JSON de volta para Excel
    json_para_excel_openpyxl("registro_openpyxl.json", "registro_openpyxl.xlsx")
    print("JSON convertido de volta para Excel com openpyxl: registro_openpyxl.xlsx")

