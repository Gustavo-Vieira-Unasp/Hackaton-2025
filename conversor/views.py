from django.shortcuts import render
import os
import json
from io import BytesIO
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from openpyxl import load_workbook, Workbook


def _get_data_dir():
    """
    Garante que existe uma pasta 'data' dentro do BASE_DIR do projeto
    e retorna o caminho absoluto dessa pasta.
    """
    data_dir = os.path.join(settings.BASE_DIR, "data")
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def _salvar_upload_excel_em_data(arquivo_uploaded):
    """
    Salva o arquivo Excel enviado pelo usuário dentro de BASE_DIR/data
    usando um nome único baseado em timestamp.

    Retorna o caminho completo do arquivo salvo.
    """
    data_dir = _get_data_dir()

    timestamp = now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"upload_{timestamp}.xlsx"

    caminho_final = os.path.join(data_dir, nome_arquivo)

    # salva em disco de forma segura (sem caminho hardcoded do PC)
    with open(caminho_final, "wb+") as destino:
        for chunk in arquivo_uploaded.chunks():
            destino.write(chunk)

    return caminho_final


def _salvar_excel_gerado_em_data(buffer_excel):
    """
    Recebe um BytesIO (planilha gerada em memória),
    salva em BASE_DIR/data com nome único,
    retorna o caminho salvo e o nome sugerido para download.
    """
    data_dir = _get_data_dir()

    timestamp = now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"gerado_{timestamp}.xlsx"
    caminho_final = os.path.join(data_dir, nome_arquivo)

    with open(caminho_final, "wb") as f:
        f.write(buffer_excel.getvalue())

    return caminho_final, nome_arquivo


def _excel_para_lista_dict(caminho_arquivo_excel):
    """
    Lê um arquivo Excel que já está salvo em disco (caminho em data/)
    e converte para lista de dicionários.
    A primeira linha da planilha é usada como cabeçalho.
    """
    wb = load_workbook(caminho_arquivo_excel, data_only=True)
    ws = wb.active

    cabecalhos = [cell.value for cell in ws[1]]

    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        linha_dict = {}
        for i, nome_coluna in enumerate(cabecalhos):
            linha_dict[nome_coluna] = row[i]
        dados.append(linha_dict)

    return dados


def _lista_dict_para_excel_bytes(lista_dict):
    """
    Recebe uma lista de dicionários (Ex.: [{"nome":"Ana","tel":"123"}, ...])
    Gera uma planilha Excel em memória (BytesIO) e devolve esse buffer.
    """
    wb = Workbook()
    ws = wb.active

    if not lista_dict:
        ws.append(["(vazio)"])
    else:
        cabecalhos = list(lista_dict[0].keys())
        ws.append(cabecalhos)

        for item in lista_dict:
            ws.append([item.get(coluna) for coluna in cabecalhos])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@csrf_exempt
def excel_para_json_view(request):
    """
    POST /upload-excel/
    Espera multipart/form-data com campo 'arquivo_excel'.

    Fluxo atualizado:
    1. Salva o arquivo Excel enviado em BASE_DIR/data.
    2. Lê esse arquivo salvo.
    3. Converte para lista de dicionários.
    4. Retorna JSON com os dados.

    Também retorna o nome do arquivo salvo em 'data' pra rastrear depois.
    """
    if request.method != "POST":
        return JsonResponse(
            {"erro": "Use método POST e envie arquivo_excel (.xlsx)."},
            status=405
        )

    if "arquivo_excel" not in request.FILES:
        return JsonResponse(
            {"erro": "Campo 'arquivo_excel' não foi enviado."},
            status=400
        )

    arquivo_enviado = request.FILES["arquivo_excel"]

    # 1. salva o upload fisicamente em BASE_DIR/data
    caminho_salvo = _salvar_upload_excel_em_data(arquivo_enviado)

    # 2. lê o conteúdo do arquivo salvo e converte em lista[dict]
    dados_convertidos = _excel_para_lista_dict(caminho_salvo)

    return JsonResponse(
        {
            "status": "ok",
            "timestamp": now().isoformat(),
            "arquivo_salvo": os.path.basename(caminho_salvo),
            "linhas_lidas": len(dados_convertidos),
            "dados": dados_convertidos,
        },
        json_dumps_params={
            "ensure_ascii": False,
            "indent": 4,
        },
        safe=False
    )


@csrf_exempt
def json_para_excel_view(request):
    """
    POST /gerar-excel/
    Espera no corpo (raw body) um JSON que seja uma lista de objetos, ex.:

    [
        {"nome": "Ana", "telefone": "5511999999999", "data_visita": "2025-10-25"},
        {"nome": "Bruno", "telefone": "5511888888888", "data_visita": "2025-10-25"}
    ]

    Fluxo atualizado:
    1. Gera um Excel em memória a partir da lista.
    2. Salva esse Excel em BASE_DIR/data.
    3. Retorna o arquivo Excel para download.
    """
    if request.method != "POST":
        return JsonResponse(
            {"erro": "Use método POST e envie um JSON no corpo da requisição."},
            status=405
        )

    try:
        corpo_bytes = request.body
        lista_dict = json.loads(corpo_bytes.decode("utf-8"))

        if not isinstance(lista_dict, list):
            raise ValueError("O corpo precisa ser uma lista JSON de objetos.")
    except Exception as e:
        return JsonResponse(
            {"erro": f"JSON inválido: {str(e)}"},
            status=400
        )

    # 1. Gera Excel em memória a partir da lista
    buffer_excel = _lista_dict_para_excel_bytes(lista_dict)

    # 2. Salva esse Excel em BASE_DIR/data
    caminho_salvo, nome_arquivo = _salvar_excel_gerado_em_data(buffer_excel)

    # 3. Retorna o Excel como download
    resposta = HttpResponse(
        buffer_excel.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resposta["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'

    # Opcionalmente, você também pode expor no header qual foi o caminho salvo
    resposta["X-Salvo-Em"] = os.path.basename(caminho_salvo)

    return resposta


def conversor_status(request):
    """
    GET /
    Só pra inspecionar rapidamente se o app está ativo e lembrar as rotas.
    """
    return HttpResponse(
        "Conversor ativo.\n\n"
        "POST /upload-excel/   -> envia arquivo_excel (.xlsx), salva em /data, retorna JSON\n"
        "POST /gerar-excel/    -> envia lista JSON, salva Excel em /data e baixa o .xlsx\n"
    )
